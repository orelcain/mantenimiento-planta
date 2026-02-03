import openpyxl, json, re

file_path = 'd:/a/APP leventamiento de insidencias en planta/Inspeccion diaria.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

result = []
current_area = None

def normalize(x):
    if x is None: return ''
    return str(x).strip()

def is_area_start(x):
    return re.match(r'^\d+\.\s+', x)

def is_ignored(x):
    ignored = ['ACTIVIDAD', 'C', 'N/C', 'OBSERVACION', 'FECHA', 'HORA', 'FOLIO', 'CODIGO', 'PAGINA', 'Página', 'REVISION', 'REGISTRO', 'RESPONSABLE', 'REVISO']
    for i in ignored:
        if i in x.upper(): return True
    return False

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 3: continue
        
        val_b = normalize(row[1])
        val_c = normalize(row[2])

        if is_area_start(val_b):
            current_area = {'name': val_b, 'items': []}
            result.append(current_area)
            continue
            
        if not current_area or is_ignored(val_b) or is_ignored(val_c):
            continue
            
        item_text = ''
        if val_c:
             if val_b and not is_area_start(val_b): 
                 item_text = f"{val_b}: {val_c}"
             else:
                 item_text = val_c
        elif val_b and not is_area_start(val_b):
             if len(val_b) > 4: 
                 item_text = val_b
        
        if item_text:
             current_area['items'].append(item_text)

print(json.dumps(result, indent=2, ensure_ascii=False))
